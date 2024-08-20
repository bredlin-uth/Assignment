import datetime
import os
from configparser import ConfigParser

import openpyxl


def get_row_count(file, sheet_name):
    workbook = openpyxl.open(file)
    sheet = workbook[sheet_name]
    return sheet.max_row

def get_column_count(file, sheet_name):
    workbook = openpyxl.open(file)
    sheet = workbook[sheet_name]
    return sheet.max_column

def read_data_from_excel(file, sheet_name, row, column):
    workbook = openpyxl.open(file)
    sheet = workbook[sheet_name]
    return sheet.cell(row, column).value

def write_data_into_excel(file, sheet_name, row, column, data):
    workbook = openpyxl.open(file)
    sheet = workbook[sheet_name]
    sheet.cell(row, column).value = data
    workbook.save(file)

def get_config(section, option):
    config = ConfigParser()
    file = os.path.join(os.path.dirname(os.path.abspath('.')), "test_data\\config.ini")
    config.read(file)
    return config.get(section, option)

def get_date(text):
    today = datetime.date.today()
    if str(text).__contains__("hour"):
        return today.strftime("%d %b %Y")
    elif str(text).__contains__("days ago"):
        digits = filter(str.isdigit, text)
        integer_str = int(''.join(digits))
        date_to_print = today - datetime.timedelta(days=integer_str)
        return date_to_print.strftime("%d %b %Y")
    elif str(text).__contains__("week ago"):
        return text
    elif str(text).__contains__("month ago"):
        return text
    elif str(text).__contains__("year ago"):
        return text
    else:
        return text
